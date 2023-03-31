#!/usr/bin/env python
# coding: utf-8

# In[17]:


# pip install python-barcode
# pip install "python-barcode[images]"
import pandas as pd
import numpy as np
import pyodbc
import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.engine import URL
from pathlib import Path
from openpyxl import load_workbook, styles, formatting
import sys
import os


# In[18]:


server = '10.1.3.25' 
database = '**'
username = '**' 
password = '**' 
connection_string = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password
connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": connection_string})
engine = create_engine(connection_url)
print("Connection Established:")


# In[19]:


df=pd.read_sql('''
WITH T1 as (
SELECT case when salesorg = '1100' Then 'IVY' ELSE 'RED' END AS [company],material, description, ip, ct, nsp, srp, upc
FROM [ivy.mm.dim.mtrl]
WHERE ivykiss = 'X' and mobile = 'X' and ms in ('01','41','91','N1','D1'))
SELECT *
FROM T1
WHERE nsp is not null and upc is not null 
''',con=engine)
df=df.astype({"company":"str","material":"str", "description":"str","ip":"int","ct":"int","nsp":"float","srp":"float","upc":"str"}).sort_values(by=["upc"],ascending=True)


# In[21]:


# Create This Week's Barcode Report
from openpyxl import load_workbook
from datetime import datetime, timedelta
import shutil
import openpyxl as op
from openpyxl import workbook
from openpyxl.styles.fonts import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter
date = datetime.today().strftime("%m%d%Y")
# Count Row Number of Dataframe
a = len(df.index)
b = len(df.columns)
count = a*b
wb = op.Workbook()
ws = wb.active
i = 0
#
ws.row_dimensions[1].height = 50
title = ws.cell(row=1,column =1, value = "IVY Beauty UPC List")
ws.sheet_view.showGridLines = False
title.font = Font(size =18, name = 'Calibri (Body)', color = "002060", bold = True )
Subtitle = ws.cell(row=2, column = 1, value = "제공되는 정보는 IVY에서 동의한  고객을 위한것이며, 본 정보의 전부 혹은 일부를 무단으로 제3자에게 공개, 배포, 복사 또는 사용하는 것은 엄격히 금지됩니다.")
Subtitle.font = Font(size = 12, name = 'Calibri (Body)', bold= True)
sub2 = ws.cell(row=3, column=1,value = "This UPC Code information is intended solely for so long as you are permitted by Ivy Beauty to use the www.IVYKISS.COM site, " )
sub2.font = Font(size=12,name = 'Calibri (Body)', bold= True)
sub3 = ws.cell(row=4,column = 1, value ="during which time you may access, view, download, and print the UPC Code information for your business use." )
sub3.font = Font(size=12,name = 'Calibri (Body)', bold= True)
sub4 = ws.cell(row=5, column = 1, value = "Please note that this UPC Code information may contain secret, privileged, or confidential information protected under applicable law. ")
sub4.font = Font(size=12,name = 'Calibri (Body)', bold= True)
sub5 = ws.cell(row= 6, column = 1, value = "Any unauthorized dissemination, prohibited copying or use of the information contained in this UPC information is strictly prohibited")
sub5.font = Font(size = 12, name = 'Calibri (Body)', bold= True)
sub6 = ws.cell(row=7, column=  1, value = "If you have any questions, please call our customer service at (516) 621-9779   Ivy Beauty")
sub6.font = Font(size = 12, name = 'Calibri (Body)', bold= True)
for col in range(1,9):
    A = ws.cell(row=1, column = col)
    A.border = Border(top=Side(border_style = None),bottom = Side(border_style = "thin"), left =  Side(border_style =None), right =  Side(border_style =None))
    for r in range(2,8):
        B = ws.cell(row=r, column = col)
        B.fill = PatternFill(fgColor = "C5D9F1", fill_type = "solid")
        B.border = Border(top = Side(border_style = None),bottom = Side(border_style = None), left =  Side(border_style =None), right =  Side(border_style =None))
for row in range(2,9):
    C = ws.cell(row = row, column = 8)
    C.border = Border(top = Side(border_style = None),bottom = Side(border_style = None), left =  Side(border_style =None), right =  Side(border_style ="thin"))
for c in range(1,9):
    C = ws.cell(row =8, column = c) 
    C.border = Border(top = Side(border_style = "thin"),bottom = Side(border_style = "thin"), left =  Side(border_style ="thin"), right =  Side(border_style ="thin"))
C = ws.cell(row = 8, column=1, value = "Company")
C.font = Font(size = 12, name = 'Calibri (Body)',bold = True)
C.alignment = Alignment(horizontal ='center',vertical = 'center')
C = ws.cell(row = 8,column = 2, value = "Material")
C.font = Font(size = 12, name = 'Calibri (Body)',bold = True) 
C.alignment = Alignment(horizontal='center',vertical='center')
C = ws.cell(row = 8, column = 3, value = "Description")
C.font = Font(size = 12, name ='Calibri (Body)',bold = True)
C.alignment = Alignment(horizontal='center',vertical = 'center')
C = ws.cell(row = 8, column = 4, value = "Inner Pack Qty")
C.font =Font(size = 12, name ='Calibri (Body)',bold = True)
C.alignment = Alignment(horizontal='center',vertical = 'center')
C = ws.cell(row=8, column=5, value = "Carton Qty")
C.font = Font(size = 12,name ='Calibri (Body)',bold = True )   
C.alignment =Alignment(horizontal='center',vertical = 'center')  
C = ws.cell(row=8, column=6, value = "NSP")
C.font = Font(size = 12,name ='Calibri (Body)',bold = True )   
C.alignment =Alignment(horizontal='center',vertical = 'center')    
C = ws.cell(row=8, column=7, value = "SRP")
C.font = Font(size = 12,name ='Calibri (Body)',bold = True )   
C.alignment =Alignment(horizontal='center',vertical = 'center') 
C = ws.cell(row=8, column=8, value = "UPC")
C.font = Font(size = 12,name ='Calibri (Body)',bold = True )   
C.alignment =Alignment(horizontal='center',vertical = 'center') 

for z in range(1,8):
    i=0
    for x in range(9,a+9):
        data = df.iloc[i,z-1]
        b = ws.cell(row=x, column = z, value=data)
        b.font = Font(size = 12, name = 'Calibri (Body)')
        b.alignment = Alignment(horizontal='center',vertical='center')
        b.border = Border(top = Side(border_style = "dotted"),bottom = Side(border_style = "dotted"), left = Side(border_style = "dotted"), right = Side(border_style = "dotted"))
        ws.row_dimensions[x].height = 26
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['B'].width = 18.56
        ws.column_dimensions['C'].width = 44.14
        ws.column_dimensions['D'].width = 18.57
        ws.column_dimensions['E'].width = 18.57
        ws.column_dimensions['F'].width = 18.57
        ws.column_dimensions['G'].width = 18.57
        ws.column_dimensions['H'].width = 34.71
        i+=1
e=0
for x in range(9,a+9):
    data = df.iloc[e,7]
    b = ws.cell(row = x, column = 8, value = data)
    b.font = Font(size = 12,name = 'Calibri (Body)')
    b.alignment = Alignment( horizontal = 'center', vertical = 'center')
    b.border = Border(top = Side(border_style = "dotted"),bottom = Side(border_style = "dotted"), left = Side(border_style = "dotted"), right = Side(border_style = "thin"))
    e+=1
D = ws['A9']
ws.freeze_panes = D
FullRange = "A8:"+get_column_letter(ws.max_column)+str(ws.max_row)
ws.auto_filter.ref = FullRange
print("New UPC List was made")
wb.save(r'C:\Users\KISS Admin\OneDrive - Kiss Products Inc\Desktop\Barcode Report Practice\UPC_LIST\UPC_'+date+".xlsx")
wb.close
    
        


# In[22]:


os.getcwd()
newpath = r'C:'
os.chdir(newpath)
import barcode
from barcode import UPCA
from barcode.writer import ImageWriter
for i,y in zip(df["material"], df["upc"]):
    with open(str(y)+".png","wb") as f:
        UPCA(str(y), writer=ImageWriter()).write(f)


# In[23]:


from numpy import size
import win32com.client
from openpyxl.worksheet.page import PageMargins
from openpyxl import Workbook
import openpyxl as op
from win32com import client
import openpyxl as op
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles.fonts import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection
from datetime import datetime
from PIL import Image as pi
from openpyxl.drawing.image import Image
import time
from openpyxl.drawing.xdr import XDRPoint2D,XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.utils.units import pixels_to_EMU,cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.descriptors.serialisable import Serialisable
date = datetime.today().strftime("%m%d%Y")
a = len(df.index)
b = len(df.columns)
count = a*b 
wb = op.Workbook()
ws = wb.worksheets[0]

i = 0
ws.row_dimensions[1].height = 50
title = ws.cell(row=1,column =3, value = "IVY Beauty UPC List")
ws.sheet_view.showGridLines = False
title.font = Font(size =18, name = 'Calibri (Body)', color = "002060", bold = True )
Subtitle = ws.cell(row=2, column = 3, value = "제공되는 정보는 IVY에서 동의한  고객을 위한것이며, 본 정보의 전부 혹은 일부를 무단으로 제3자에게 공개, 배포, 복사 또는 사용하는 것은 엄격히 금지됩니다.")
Subtitle.font = Font(size = 11, name = 'Calibri (Body)', bold= True)
sub2 = ws.cell(row=3, column=3,value = "This UPC Code information is intended solely for so long as you are permitted by Ivy Beauty to use the www.IVYKISS.COM site, " )
sub2.font = Font(size=11,name = 'Calibri (Body)', bold= True)
sub3 = ws.cell(row=4,column = 3, value ="during which time you may access, view, download, and print the UPC Code information for your business use.")
sub3.font = Font(size=11,name = 'Calibri (Body)', bold= True)
sub4 = ws.cell(row=5, column = 3, value = "Please note that this UPC Code information may contain secret, privileged, or confidential information protected under applicable law. ")
sub4.font = Font(size=11,name = 'Calibri (Body)', bold= True)
sub5 = ws.cell(row= 6, column = 3, value = "Any unauthorized dissemination, prohibited copying or use of the information contained in this UPC information is strictly prohibited")
sub5.font = Font(size=11,name = 'Calibri (Body)', bold = True)
sub6 = ws.cell(row=7, column=  3, value = "If you have any questions, please call our customer service at (516) 621-9779   Ivy Beauty")
sub6.font = Font(size=11,name = 'Calibri (Body)', bold = True)
A = ws.cell(row = 1,column = 9)
A.border = Border(bottom = Side(border_style="thin"))
A = ws.cell(row = 1, column = 10)
A.border = Border(bottom = Side(border_style="thin"))
for col in range(1,9):
    A = ws.cell(row=1, column = col)
    A.border = Border(top = Side(border_style =None), bottom = Side(border_style = "thin"), left =  Side(border_style =None), right =  Side(border_style =None))
    for r in range(2,8):
        B = ws.cell(row = r, column = col)
        B.fill = PatternFill(fgColor = "C5D9F1", fill_type = "solid")
        B.border = Border(top = Side(border_style = None),bottom = Side(border_style = None), left =  Side(border_style =None), right =  Side(border_style =None))
for col in range(8,11):
    for r in range(2,8):
        B = ws.cell(row = r, column = col)
        B.fill = PatternFill(fgColor = "FDE9D9", fill_type = "solid")
        B.border = Border(top = Side(border_style = None),bottom = Side(border_style = None), left =  Side(border_style =None), right =  Side(border_style =None))
ws.cell(row=2,column = 8).value = "*Ivy Scan App을 사용하시는 분들은, 제품을 스캔할 때 PDF"
ws.cell(row=3, column=8).value = "프로그램 화면 비율을 400%로 확대하여 사용 부탁드립니다."
ws.cell(row=4,column=8).value = " *For Ivy Scan App users, please magnify your PDF"
ws.cell(row=5,column=8).value = "reader zoom level to 400% when you scan the products."
ws.cell(row=2, column = 8).font = Font(size = 10, name = 'Calibri (Body)',bold = True) 
ws.cell(row=3, column=8).font =  Font(size = 10, name = 'Calibri (Body)',bold = True)
ws.cell(row=4,column=8).font =  Font(size = 10, name = 'Calibri (Body)',bold = True)
ws.cell(row=5,column=8).font =  Font(size = 10, name = 'Calibri (Body)',bold = True)
a = ws.merge_cells(start_row=6,start_column=8,end_row=7,end_column=9)
for row in range(2,8):
    C = ws.cell(row=row,column = 7)
    C.border = Border(right = Side(border_style='thin'))
for row in range(6,8):
    D = ws.cell(row=row, column=9)
    D.border=Border(left = Side(border_style=None))
for row in range(2,8):
    D = ws.cell(row = row, column = 8)
    D.border = Border(right=Side(border_style="thin"))
D = ws.cell(row = 2, column = 11)
D.border = Border(left=Side(border_style='thin'))   
a = len(df.index)
for row in range(2,10):
      C = ws.cell(row = row, column = 10)
      C.border = Border(top = Side(border_style = None),bottom = Side(border_style = None), left =  Side(border_style =None), right =  Side(border_style ="thin"))
for c in range(1, 11):
    C = ws.cell(row = 8, column = c)
    C.border =Border(top = Side(border_style = "thin"),bottom = Side(border_style = "thin"), left =  Side(border_style ="thin"), right =  Side(border_style ="thin"))
C = ws.cell(row = 8, column = 1, value = "MATERIAL")
C.font = Font(size = 12, name = 'Calibri (Body)', bold = True)
C.alignment = Alignment(horizontal = 'center',vertical = 'center')
C = ws.cell(row = 8, column=2, value = "DESCRIPTION")
C.font = Font(size = 12, name = 'Calibri (Body)', bold = True)
C.alignment = Alignment(horizontal = 'center',vertical = 'center')
C = ws.cell(row = 8, column = 3, value = "COMPANY")
C.font = Font(size = 12, name = 'Calibri (Body)', bold = True)
C.alignment = Alignment(horizontal = 'center',vertical = 'center')
C = ws.cell(row = 8, column = 4, value = "MATERIAL")
C.font = Font(size = 12, name = 'Calibri (Body)', bold = True)
C.alignment = Alignment(horizontal = 'center',vertical = 'center')
C = ws.cell(row = 8, column = 5,value = "BARCODE")
C.font = Font(size = 12, name = 'Calibri (Body)', bold = True)
C.alignment = Alignment(horizontal = 'center',vertical = 'center')
C = ws.cell(row = 8, column = 6, value = "Inner Pack Qty")
C.font = Font(size = 12, name = 'Calibri (Body)', bold = True)
C.alignment = Alignment(horizontal = 'center',vertical = 'center')
C = ws.cell(row = 8, column = 7, value = "Carton Qty")
C.font = Font(size = 12, name = 'Calibri (Body)', bold = True)
C.alignment = Alignment(horizontal = 'center',vertical = 'center')
C = ws.cell(row = 8, column = 8, value = "NSP")
C.font = Font(size = 12, name = 'Calibri (Body)', bold = True)
C.alignment = Alignment(horizontal = 'center',vertical = 'center')
C = ws.cell(row =8, column = 9,value = "SRP")
C.font = Font(size = 12, name = 'Calibri (Body)', bold = True)
C.alignment = Alignment(horizontal = 'center',vertical = 'center')
C = ws.cell(row = 8, column = 10, value = "UPC")
C.font = Font(size = 12, name = 'Calibri (Body)', bold = True)
C.alignment = Alignment(horizontal = 'center',vertical = 'center')
# 컬럼 AB에 형식과 데이터를 넣기 
for z in range(2,4):
    i=0
    for x in range(9,a+9):
        data = df.iloc[i,z-1]
        b = ws.cell(row = x, column = z-1,value = data)
        b.font = Font(size =11, name = 'Calibri (Body)')
        b.alignment = Alignment(horizontal = 'center',vertical = 'center')
        b.border = Border(top = Side(border_style = "thin"),bottom = Side(border_style = "thin"), left = Side(border_style = "thin"), right = Side(border_style = "thin"))
        ws.row_dimensions[x].height = 150
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 53
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 44
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 20
        i+=1
# 칼럼 C에 형식과 데이터를 넣기 
for x in range(9,a+9):
    i=0
    data = df.iloc[i,0]
    b = ws.cell(row = x,column = 3,value = data)
    b.font = Font(size = 12, name = 'Calibri (Body)')
    b.alignment = Alignment(horizontal='center',vertical='center')
    b.border = Border(top = Side(border_style = "thin"),bottom = Side(border_style = "thin"), left = Side(border_style = "thin"), right = Side(border_style = "thin"))
    i+=1
# 칼럼 D에는 A와 B를 텍스트 concat해서 넣기 
for x in range(9,a+9):
    data = df.iloc[x-9,2]
    data2 = df.iloc[x-9,1]
    b = ws.cell(row = x, column = 4, value = data +'(' + str(data2) + ')')
    b.font = Font(size = 11, name = 'Calibri (Body)')
    b.alignment = Alignment(horizontal='center',vertical='center')
    b.border =  Border(top = Side(border_style = "thin"),bottom = Side(border_style = "thin"), left = Side(border_style = "thin"), right = Side(border_style = "thin"))
#F/G/H/I/J column에 데이터와 형식을 넣기 
for z in range(6, 11):
    i=0
    for x in range(9,a+9):
        data = df.iloc[i,z-3]
        b = ws.cell(row = x,column = z, value = data)
        b.alignment = Alignment(horizontal='center',vertical = 'center')
        b.font = Font(size=11,name = 'Calibri (Body)')
        i+=1
for h in range(8,10):
    for x in range(9,a+9):
        b = ws.cell(row = x,column = h)
        b.number_format = u'$#,##0.00;'
        
for x in range(9,11):
    b = ws.cell(row = 2, column = x)
    b.border = Border(top=Side(border_style='thin'))
# 9번 row부터 border line 넣기
for r in range(9,a+9):
    for col in range(1,11):
        D = ws.cell(row = r, column = col)
        D.border = Border(top = Side(border_style = "thin"),bottom = Side(border_style = "thin"), left = Side(border_style = "thin"), right = Side(border_style = "thin"))

#E column에 바코드 이미지 자동으로 넣기 
no_image = []
for x in range(9,a+9):
    b = ws.cell(row=x, column=5)
    img_path= r'C:\Users\KISS Admin\OneDrive - Kiss Products Inc\Desktop\Barcode Report Practice\Barcode_by_python'
    name = ws.cell(row=x,column=10).value
    try:
        img = "C:\\Users\\KISS Admin\\OneDrive - Kiss Products Inc\\Desktop\\Barcode Report Practice\\Barcode_by_python\\"+name+".png"
        img2=op.drawing.image.Image(img)
        img2.height = 300
        img2.width=150
        c2e = cm_to_EMU
        p2e = pixels_to_EMU
        size = XDRPositiveSize2D(p2e(img2.height),p2e(img2.width))
        cellh = lambda x: c2e((x *400)/99)
        cellw = lambda x: c2e((x * (18.65 - 10)/10))
        column=5;row = x;coloffset = cellw(0.2);rowoffset=cellh(0.2)
        marker = AnchorMarker(col = column-1, colOff = coloffset, row = row-1, rowOff = rowoffset)
        img2.anchor =  OneCellAnchor(_from = marker, ext= size)
        ws.add_image(img2)
        b = ws.cell(row=x,column=5)
        x+=1
    except FileNotFoundError:
        no_image.appned(name)
        x+=1                   

D = ws['A9']
ws.freeze_panes = D
FullRange = "A8:"+get_column_letter(ws.max_column)+str(ws.max_row)
ws.auto_filter.ref = FullRange

date = datetime.today().strftime("%m%d%Y")
wb.save(r'C:\Users\KISS Admin\OneDrive - Kiss Products Inc\Desktop\Barcode Report Practice\Barcode_Report\Barcode_' + date+ '.xlsx')
wb.close()
print(no_image)

